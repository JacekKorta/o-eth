from os.path import join
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import Settings


def copy_column(translation_file_name: str, column_name: str, settings: Settings) -> None:
    result_workbook = load_workbook(filename=settings.result_filename)
    translation_workbook = load_workbook(filename=join(settings.work_dir, translation_file_name))

    result_sheet = result_workbook.active
    translation_sheet = translation_workbook.active

    translation_col = translation_sheet[settings.translation_column_name]

    for cell in translation_col:
        if cell.row > settings.header_row_number:
            result_sheet[f"{column_name}{cell.row}"] = cell.value

    result_workbook.save(filename=settings.result_filename)


def find_column(header_value: str, settings: Settings) -> Optional[str]:
    result_workbook = load_workbook(filename=settings.result_filename)
    result_sheet = result_workbook.active

    header_tuple = result_sheet[str(settings.header_row_number)]

    for cell in header_tuple:
        if cell.value == header_value:
            return get_column_letter(cell.column)
    #  TODO: placeholder for fixing source data
    return None
