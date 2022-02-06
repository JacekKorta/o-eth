from os.path import join

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import get_settings

settings = get_settings()


def copy_column(
    translation_file_name: str, result_file_name: str, column_name: str, main_dir: str
) -> None:
    result_workbook = load_workbook(filename=result_file_name)
    translation_workbook = load_workbook(filename=join(main_dir, translation_file_name))

    result_sheet = result_workbook.active
    translation_sheet = translation_workbook.active

    translation_col = translation_sheet[settings.translation_column_name]

    for cell in translation_col:
        if cell.row > settings.header_row_number:
            result_sheet[f"{column_name}{cell.row}"] = cell.value

    result_workbook.save(filename=result_file_name)


def find_column(result_filename: str, header_value: str) -> str:
    result_workbook = load_workbook(filename=result_filename)
    result_sheet = result_workbook.active

    header_tuple = result_sheet[str(settings.header_row_number)]

    for cell in header_tuple:
        if cell.value == header_value:
            return get_column_letter(cell.column)
